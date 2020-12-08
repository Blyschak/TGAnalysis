import os
import tempfile
import tkinter as tk
from dataclasses import dataclass
from tkinter import filedialog, END
from tkinter import messagebox
from typing import Iterable

import matplotlib.backends.backend_tkagg as tkagg
import matplotlib.figure as mplfig
import openpyxl as openpyxl
from matplotlib.backends._backend_tk import NavigationToolbar2Tk
from openpyxl.utils import column_index_from_string

import openfile
import tga
from openfile import openfile


@dataclass(frozen=True)
class TemperatureRange:
    T_first: float
    T_second: float


class TemperatureInputFrame(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.temperature_input = tk.Entry(self)
        self.temperature_label = tk.Label(self, text="T, C")
        self.temperature_label.grid(row=0, column=1)
        self.temperature_input.grid(row=0, column=2)

    def get_value(self):
        return float(self.temperature_input.get())


class TemperatureRangeInputFrame(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.temperature_inputs = []

        self.temperature_input_container = tk.Frame(self)

        self.temperature_input_add_btn = tk.Button(self,
                                                   text="Add Temperature Point",
                                                   command=self.add_temperature_input)
        self.temperature_input_discard_btn = tk.Button(self,
                                                       text="Discard",
                                                       command=self.discard_temperature_input)

        self.temperature_input_add_btn.grid(row=0, column=0)
        self.temperature_input_discard_btn.grid(row=0, column=1)
        self.temperature_input_container.grid(row=1, columnspan=2)

        self.add_temperature_input()
        self.add_temperature_input()

    def add_temperature_input(self):
        temperature_input = TemperatureInputFrame(self.temperature_input_container)
        temperature_input.pack()
        self.temperature_inputs.append(temperature_input)

    def discard_temperature_input(self):
        if len(self.temperature_inputs) < 3:
            return
        temperature_input = self.temperature_inputs.pop()
        temperature_input.destroy()

    def get_temperature_ranges(self) -> Iterable[TemperatureRange]:
        for temp_inputs in zip(self.temperature_inputs, self.temperature_inputs[1:]):
            first, second = temp_inputs
            yield TemperatureRange(first.get_value(), second.get_value())


class ResultsExporter:
    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

    def export(self, range_results):
        row = 1
        for temperature_range, results in range_results.items():
            temp_first, temp_second = temperature_range.T_first, temperature_range.T_second
            self.sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=6)
            self.sheet.cell(row=row, column=1).value = f'T: [{temp_first}C-{temp_second}C]'
            row += 1
            self.sheet.cell(row=row, column=1).value = 'Model'
            self.sheet.cell(row=row, column=2).value = 'E_a, kJ/mol'
            self.sheet.cell(row=row, column=3).value = 'delta(E_a), kJ/mol'
            self.sheet.cell(row=row, column=4).value = 'lnA'
            self.sheet.cell(row=row, column=5).value = 'delta(lnA)'
            self.sheet.cell(row=row, column=6).value = 'r_value'
            row += 1
            for idx, (model, params) in enumerate(results.items()):
                self.sheet.cell(row=row, column=1).value = model
                self.sheet.cell(row=row, column=2).value = round(params.Ea / 1000., 1)
                self.sheet.cell(row=row, column=3).value = round(params.delta_Ea / 1000., 1)
                self.sheet.cell(row=row, column=4).value = round(params.lnA, 1)
                self.sheet.cell(row=row, column=5).value = round(params.delta_lnA, 1)
                self.sheet.cell(row=row, column=6).value = round(params.rvalue, 3)
                row += 1

        self.wb.save(self.filepath)


class XlsTGALoader:
    def __init__(self, filepath,
                 temp_column='A',
                 alpha_column='B',
                 derivative_column='C',
                 sheet_name=None,
                 units='C',
                 heat_rate=1.):
        self.filepath = filepath
        self.temp_column = column_index_from_string(temp_column)
        self.alpha_column = column_index_from_string(alpha_column)
        self.derivative_column = column_index_from_string(derivative_column)
        self.sheet_name = sheet_name
        self.units = units
        self.heat_rate = heat_rate

    def get_tga(self):
        wb = openpyxl.load_workbook(self.filepath)
        sheet = wb.active
        if self.sheet_name is not None:
            sheet = wb.get_sheet_by_name(self.sheet_name)

        temps = []
        alphas = []
        derivs = []
        for i in range(1, sheet.max_row + 1):
            temp_string = sheet.cell(i, self.temp_column).value
            alpha_string = sheet.cell(i, self.alpha_column).value
            deriv_string = sheet.cell(i, self.derivative_column).value

            # assume no more data
            if not temp_string:
                break

            try:
                temps.append(float(temp_string))
            except ValueError as err:
                raise ValueError(f'Invalid number at {i}:{self.temp_column}: {err}')

            try:
                alphas.append(float(alpha_string))
            except ValueError as err:
                raise ValueError(f'Invalid number at {i}:{self.temp_column}: {err}')

            try:
                derivs.append(float(deriv_string))
            except ValueError as err:
                raise ValueError(f'Invalid number at {i}:{self.temp_column}: {err}')

        return tga.TGA(temps, alphas, derivs,
                       t_units=self.units,
                       heating_rate=self.heat_rate)


class MainWindow(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.tga = None

        self.temp_range_input = TemperatureRangeInputFrame(self)
        self.temp_range_input.grid(row=0, column=0)

        self.load_xlsx_data_btn = tk.Button(self, text="Load xlsx with TGA data",
                                            command=self.load_from_xlsx)
        self.load_xlsx_data_btn.grid(row=1, column=0)

        self.heat_rate_label = tk.Label(self, text="Heat Rate: (K/min)")
        self.heat_rate_label.grid(row=1, column=1)

        self.heat_rate_entry = tk.Entry(self, text="Heat Rate: (K/min)")
        self.heat_rate_entry.insert(END, "10")
        self.heat_rate_entry.grid(row=1, column=2)

        self.temp_column_label = tk.Label(self, text="Temperature Column:")
        self.temp_column_label.grid(row=2, column=0)
        self.temp_column_entry = tk.Entry(self, text="Temperature Column:")
        self.temp_column_entry.insert(END, "A")
        self.temp_column_entry.grid(row=2, column=1)

        self.alpha_column_label = tk.Label(self, text="Alpha Column:")
        self.alpha_column_label.grid(row=3, column=0)
        self.alpha_column_entry = tk.Entry(self, text="Alpha Column:")
        self.alpha_column_entry.insert(END, "B")
        self.alpha_column_entry.grid(row=3, column=1)

        self.deriv_column_lable = tk.Label(self, text="DTG Column:")
        self.deriv_column_lable.grid(row=4, column=0)
        self.deriv_column_entry = tk.Entry(self, text="DTG Column:")
        self.deriv_column_entry.insert(END, "C")
        self.deriv_column_entry.grid(row=4, column=1)


        self.fit_model_btn = tk.Button(self, text="Fit kinetic model",
                                       command=self.fit_model)
        self.fit_model_btn.grid(row=1, column=3)

        self.fig = mplfig.Figure(figsize=(5, 5))
        self.canvas = tkagg.FigureCanvasTkAgg(self.fig, master=self)

        self.draw_tga()

    def get_heat_rate(self):
        try:
            return float(self.heat_rate_entry.get()) / 60.
        except ValueError as err:
            raise ValueError(f'Invalid heating rate: {err}')

    def load_from_xlsx(self):
        try:
            filepath = filedialog.askopenfilename()
            if not filepath:
                return
            loader = XlsTGALoader(filepath,
                                  temp_column=self.temp_column_entry.get(),
                                  alpha_column=self.alpha_column_entry.get(),
                                  heat_rate=self.get_heat_rate())
            self.tga = loader.get_tga()
            self.draw_tga()
        except Exception as err:
            messagebox.showerror("Error", f"Invalid data input {err}")

    def fit_model(self):
        if self.tga is None:
            return

        model_fit = tga.ModelFitting()
        try:
            temperature_ranges = self.temp_range_input.get_temperature_ranges()
            temperature_ranges_results = {}
            for trange in temperature_ranges:
                tga_range = self.tga.range(trange.T_first, trange.T_second)
                results = model_fit.fit(tga_range)
                temperature_ranges_results[trange] = results

            tempfilepath = os.path.join(tempfile.gettempdir(),
                                        tempfile.gettempprefix() + '.xlsx')
            ResultsExporter(tempfilepath).export(temperature_ranges_results)
            openfile(tempfilepath)
        except Exception as err:
            messagebox.showerror("Error", f"Error: {err}")

    def draw_tga(self):
        self.fig.clf()
        plot = self.fig.subplots()
        if self.tga is not None:
            plot.plot(self.tga.temp, self.tga.alpha, label="TGA curve")
            plot.set_xlabel("T, C")
            plot.set_ylabel("alpha, %")
            plot.legend()

            dt_plot = plot.twinx()
            dt_plot.plot(self.tga.temp, self.tga.dtg,
                         label="DTG curve", color='red')
            dt_plot.set_ylabel("d(alpha)/dT")
            dt_plot.legend()

        self.canvas = tkagg.FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().grid(row=0, column=5)

        self.canvas.draw()

        toolbarFrame = tk.Frame(master=self)
        toolbar = NavigationToolbar2Tk(self.canvas,
                                       toolbarFrame)
        toolbar.update()
        toolbarFrame.grid(row=1, column=5)


def main():
    root = tk.Tk()
    app = MainWindow(root)
    app.pack()
    tk.mainloop()


if __name__ == '__main__':
    main()